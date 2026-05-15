"""XLSX Content Extractor"""

from pathlib import Path


class XlsxExtractor:
    """Excel document content extraction"""

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}

    def extract_text(self, file_path: str | Path) -> str:
        """Extract text content from an XLSX file (all sheets)"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return self._fallback_extract(file_path)
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            try:
                parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    parts.append(f"=== {sheet_name} ===")
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) for c in row if c is not None]
                        if cells:
                            parts.append(" | ".join(cells))
                return "\n".join(parts)
            finally:
                wb.close()
        except Exception:
            return self._fallback_extract(file_path)

    def _fallback_extract(self, file_path: str | Path) -> str:
        """Fallback when openpyxl is unavailable — returns empty since xlsx is a ZIP format."""
        return ""

    def extract_metadata(self, file_path: str | Path) -> dict:
        """Extract XLSX metadata"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {}
        try:
            wb = load_workbook(str(file_path), read_only=True)
            try:
                meta = {
                    "sheets": wb.sheetnames,
                    "sheet_count": len(wb.sheetnames),
                }
                if wb.properties:
                    meta["title"] = wb.properties.title or ""
                    meta["author"] = wb.properties.creator or ""
                return meta
            finally:
                wb.close()
        except Exception:
            return {}
